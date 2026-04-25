#!/usr/bin/env python3
"""
health_system_lookup.py

Classify providers by health system using a three-phase approach:

  Phase 1  — Reference file lookup (name match → address match → addr1 org detection)
  Phase 1b — Provider name Google Places search (disambiguates same-address/diff-org cases)
  Phase 2  — Google Places API for remaining unique unclassified addresses
  Phase 3  — Write results + summary table

Populates columns:
  health_system      (col O)
  health_system_city (col P)

API key:
  Set environment variable before running:
    export GOOGLE_PLACES_API_KEY="your-key-here"

Usage:
    python3 health_system_lookup.py
"""

import os
import re
import sys
import time
from pathlib import Path

import requests
import pandas as pd
import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────
BASE        = Path.home() / "Downloads" / "CRC MDH Project"
INPUT_FILE  = BASE / "ProviderDataFiles" / "Mailing List for Printing Services copy_columns_corrected_orgs.xlsx"
REF_FILE    = BASE / "ProviderDataFiles" / "GHmatched_phyPA_LMH.xlsx"
OUTPUT_FILE = BASE / "ProviderDataFiles" / "Mailing List for Printing Services copy_columns_corrected_orgs_hs.xlsx"

# Column indices (0-based) for health_system and health_system_city in input file
COL_HEALTH_SYSTEM      = 14   # column O
COL_HEALTH_SYSTEM_CITY = 15   # column P

# ── API config ───────────────────────────────────────────────────────────────
API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "")
PLACES_URL = "https://places.googleapis.com/v1/places:searchText"
API_DELAY  = 0.15   # seconds between API calls

# ── Classification patterns ──────────────────────────────────────────────────
MEDICAL_RE = re.compile(
    r"clinic|hospital|medical|health|care|wellness|rehab|therapy|therapist|"
    r"surgery|surgical|ortho|cardio|oncol|pediatr|neuro|psych|mental|dental|"
    r"vision|eye|family practice|urgent care|emergency|pharmacy|pharma|"
    r"clinica|salud|centro|community|associates|group|practice|services|urgent",
    re.IGNORECASE
)

REALESTATE_RE = re.compile(
    r"airbnb|vrbo|zillow|trulia|redfin|realtor|keller williams|coldwell|"
    r"century 21|re/max|remax|sotheby|real estate|realty|properties|"
    r"apartments|apartment|condo|rental|rentals|suites|inn|hotel|motel|"
    r"lodge|resort|vacation",
    re.IGNORECASE
)

# ── Helpers ──────────────────────────────────────────────────────────────────
def norm(s) -> str:
    if pd.isna(s) or s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).lower().strip())

def find_col(df, candidates):
    low = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c in df.columns:
            return c
        if c.lower() in low:
            return low[c.lower()]
    return None

def is_po_box(addr: str) -> bool:
    return bool(re.match(r"^\s*p\.?\s*o\.?\s*box\b", addr, re.IGNORECASE))

def is_org_name(addr1: str) -> bool:
    """Return True if addr1 looks like an org name (starts with letter, not PO Box)."""
    if not addr1:
        return False
    if is_po_box(addr1):
        return False
    return bool(re.match(r"^[a-zA-Z]", addr1.strip()))

# ── Google Places API call ───────────────────────────────────────────────────
def places_search(query: str) -> dict | None:
    """Search Google Places API v1. Returns first result dict or None."""
    if not API_KEY:
        return None
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": API_KEY,
        "X-Goog-FieldMask": "places.displayName,places.formattedAddress,places.types",
    }
    payload = {"textQuery": query, "maxResultCount": 1}
    try:
        resp = requests.post(PLACES_URL, json=payload, headers=headers, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        places = data.get("places", [])
        return places[0] if places else None
    except Exception as e:
        print(f"    API error for '{query}': {e}")
        return None

def classify_place(place: dict) -> tuple[str, str]:
    """Return (health_system_name, place_type) from a Places API result."""
    if not place:
        return "", ""
    name    = place.get("displayName", {}).get("text", "")
    address = place.get("formattedAddress", "")
    types   = place.get("types", [])

    # Extract city from formatted address (last part before country)
    city = ""
    parts = address.split(",")
    if len(parts) >= 2:
        city = parts[-3].strip() if len(parts) >= 3 else parts[-2].strip()

    if MEDICAL_RE.search(name):
        return name, city
    if REALESTATE_RE.search(name):
        return "REAL_ESTATE", city
    if any(t in ("health", "hospital", "doctor", "pharmacy") for t in types):
        return name, city
    return "", city

# ── Load reference file ───────────────────────────────────────────────────────
print(f"\nLoading reference file: {REF_FILE.name}")
try:
    ref_df = pd.read_excel(REF_FILE, dtype=str)
except FileNotFoundError:
    print(f"  WARNING: reference file not found, skipping Phase 1 lookup")
    ref_df = None

ref_lookup = {}   # norm_addr → (health_system, health_system_city)
if ref_df is not None:
    hs_col   = find_col(ref_df, ["Health System Name"])
    city_col = find_col(ref_df, ["Health System City"])
    for addr_col in ["work_address_1", "mailing_address_1", "home_address_1", "primary_address_1"]:
        col = find_col(ref_df, [addr_col])
        if not col:
            continue
        for _, row in ref_df.iterrows():
            hs   = str(row[hs_col]).strip()  if hs_col   and not pd.isna(row.get(hs_col, ""))   else ""
            city = str(row[city_col]).strip() if city_col and not pd.isna(row.get(city_col, "")) else ""
            if hs:
                key = norm(row[col])
                if key and key not in ref_lookup:
                    ref_lookup[key] = (hs, city)
    print(f"  {len(ref_lookup)} address entries in reference lookup")

# ── Load input file with openpyxl (preserves blank-header columns) ────────────
print(f"\nLoading input file: {INPUT_FILE.name}")
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active

headers = [str(cell.value) if cell.value is not None else f"__col{i}__"
           for i, cell in enumerate(ws[1])]

rows_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows_data.append(list(row))

df = pd.DataFrame(rows_data, columns=headers)
print(f"  {len(df)} rows | {len(df.columns)} columns")

# Detect key columns
last_col   = find_col(df, ["last_name", "LastName"])
first_col  = find_col(df, ["first_name", "FirstName"])
addr1_col  = find_col(df, ["primary_address_1", "work_address_1", "Alternate 1 Address"])
addr2_col  = find_col(df, ["primary_address_2", "work_address_2"])
addr3_col  = find_col(df, ["address_line3", "primary_address_3"])
city_col   = find_col(df, ["primary_city", "work_city", "City"])
zip_col    = find_col(df, ["zip5", "primary_postal_code", "ZIP+4"])
hs_col     = headers[COL_HEALTH_SYSTEM]
hs_city_col = headers[COL_HEALTH_SYSTEM_CITY]

print(f"  Name: last={last_col} first={first_col}")
print(f"  Addr: addr1={addr1_col} city={city_col} zip={zip_col}")
print(f"  Output cols: hs={hs_col} hs_city={hs_city_col}")

# ── Phase 1: Reference file lookup ───────────────────────────────────────────
print(f"\nPhase 1: Reference file lookup ...")
phase1_hits = 0

for i, row in df.iterrows():
    # Skip if already classified
    if norm(row.get(hs_col, "")):
        continue

    addr1 = str(row.get(addr1_col, "")).strip() if addr1_col else ""
    addr2 = str(row.get(addr2_col, "")).strip() if addr2_col else ""
    addr3 = str(row.get(addr3_col, "")).strip() if addr3_col else ""
    city  = str(row.get(city_col,  "")).strip() if city_col  else ""

    # Step 1a: addr1 org detection
    if is_org_name(addr1):
        df.at[i, hs_col]      = addr1
        df.at[i, hs_city_col] = city
        phase1_hits += 1
        continue

    # Step 1b: addr2/addr3 org detection
    for extra in [addr2, addr3]:
        if is_org_name(extra):
            df.at[i, hs_col]      = extra
            df.at[i, hs_city_col] = city
            phase1_hits += 1
            break
    if norm(df.at[i, hs_col]):
        continue

    # Step 1c: reference file address lookup
    for addr in [addr1, addr2, addr3]:
        key = norm(addr)
        if key and key in ref_lookup:
            hs, hs_city = ref_lookup[key]
            df.at[i, hs_col]      = hs
            df.at[i, hs_city_col] = hs_city or city
            phase1_hits += 1
            break

print(f"  Phase 1 classified: {phase1_hits}")
unclassified_after_p1 = df[df[hs_col].apply(norm) == ""].shape[0]
print(f"  Still unclassified: {unclassified_after_p1}")

# ── Phase 1b: Provider name search ───────────────────────────────────────────
# Searches by provider name + city for providers still unclassified after Phase 1.
# Catches cases where multiple orgs share one address (e.g., two different
# clinics in the same building) — address-based lookup can't disambiguate these.
if not API_KEY:
    print(f"\nPhase 1b: Skipped (GOOGLE_PLACES_API_KEY not set)")
else:
    unclassified_mask_1b = df[hs_col].apply(norm) == ""
    p1b_rows = df[unclassified_mask_1b]
    print(f"\nPhase 1b: Name-based search for {len(p1b_rows)} unclassified providers ...")
    phase1b_hits = 0

    # Detect specialty/taxonomy column if present
    taxonomy_col = find_col(df, ["taxonomy_descriptions", "primary_taxonomy_code",
                                  "license_type_desc", "specialty boards"])

    for i, row in p1b_rows.iterrows():
        first = str(row.get(first_col, "")).strip() if first_col else ""
        last  = str(row.get(last_col,  "")).strip() if last_col  else ""
        city  = str(row.get(city_col,  "")).strip() if city_col  else ""

        if not last:
            continue

        # Build query: name + optional specialty + city + "physician"
        specialty = ""
        if taxonomy_col:
            tax_val = str(row.get(taxonomy_col, "")).strip()
            if tax_val and tax_val.lower() not in ("nan", ""):
                # Take first word of taxonomy to keep query short
                specialty = tax_val.split()[0] if tax_val else ""

        parts = [p for p in [first, last, specialty, city, "physician"] if p]
        query = " ".join(parts)

        place = places_search(query)
        hs, hs_city = classify_place(place)

        if hs and hs != "REAL_ESTATE":
            df.at[i, hs_col]      = hs
            df.at[i, hs_city_col] = hs_city or city
            phase1b_hits += 1

        time.sleep(API_DELAY)

    print(f"  Phase 1b classified: {phase1b_hits}")
    print(f"  Still unclassified : {df[df[hs_col].apply(norm) == ''].shape[0]}")

# ── Phase 2: Google Places API for unique unclassified addresses ──────────────
if not API_KEY:
    print(f"\nPhase 2: Skipped (GOOGLE_PLACES_API_KEY not set)")
else:
    unclassified_mask = df[hs_col].apply(norm) == ""

    # Collect unique addresses to minimize API calls
    unique_addrs: dict[str, tuple] = {}  # norm_addr_key → (addr1, city, zip)
    for _, row in df[unclassified_mask].iterrows():
        addr1 = str(row.get(addr1_col, "")).strip() if addr1_col else ""
        city  = str(row.get(city_col,  "")).strip() if city_col  else ""
        z     = str(row.get(zip_col,   "")).strip() if zip_col   else ""
        key   = norm(addr1) + "|" + norm(city)
        if key not in unique_addrs:
            unique_addrs[key] = (addr1, city, z)

    print(f"\nPhase 2: API lookup for {len(unique_addrs)} unique addresses ...")
    addr_results: dict[str, tuple] = {}  # norm key → (health_system, hs_city)

    for idx, (key, (addr1, city, z)) in enumerate(unique_addrs.items()):
        if is_po_box(addr1):
            query = f"clinic hospital {city} {z}"
        else:
            query = f"{addr1} {city} {z}"

        place = places_search(query)
        hs, hs_city = classify_place(place)
        addr_results[key] = (hs, hs_city or city)

        if (idx + 1) % 50 == 0:
            print(f"  {idx + 1}/{len(unique_addrs)} addresses processed ...")
        time.sleep(API_DELAY)

    # Apply API results
    phase2_hits = 0
    for i, row in df[unclassified_mask].iterrows():
        addr1 = str(row.get(addr1_col, "")).strip() if addr1_col else ""
        city  = str(row.get(city_col,  "")).strip() if city_col  else ""
        key   = norm(addr1) + "|" + norm(city)
        hs, hs_city = addr_results.get(key, ("", ""))
        if hs and hs != "REAL_ESTATE":
            df.at[i, hs_col]      = hs
            df.at[i, hs_city_col] = hs_city
            phase2_hits += 1

    print(f"  Phase 2 classified: {phase2_hits}")
    print(f"  Still unclassified: {df[df[hs_col].apply(norm) == ''].shape[0]}")

# ── Phase 3: Write results ────────────────────────────────────────────────────
print(f"\nPhase 3: Writing results ...")

# Write back to openpyxl workbook to preserve all original columns
hs_col_idx      = COL_HEALTH_SYSTEM + 1      # openpyxl is 1-based
hs_city_col_idx = COL_HEALTH_SYSTEM_CITY + 1

for i, row in df.iterrows():
    ws.cell(row=i + 2, column=hs_col_idx,      value=row[hs_col])
    ws.cell(row=i + 2, column=hs_city_col_idx, value=row[hs_city_col])

wb.save(OUTPUT_FILE)
print(f"  Saved: {OUTPUT_FILE.name}")

# Summary
classified   = df[df[hs_col].apply(norm) != ""].shape[0]
unclassified = df[df[hs_col].apply(norm) == ""].shape[0]
print(f"\nSummary:")
print(f"  Total rows       : {len(df)}")
print(f"  Classified       : {classified}")
print(f"  Unclassified     : {unclassified}")

top_systems = df[df[hs_col].apply(norm) != ""][hs_col].value_counts().head(20)
print(f"\n  Top 20 health systems:")
for hs, cnt in top_systems.items():
    print(f"    {cnt:4d}  {hs}")

print("\nDone.")
